from io import StringIO
import os
import time
import re
import datetime as dt

import pandas as pd
import language_tool_python as lt

from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfparser import PDFParser

from tkinter.filedialog import askdirectory
from openpyxl import Workbook


# ------------------------------------------
# CONFIGURAÇÃO INICIAL
# ------------------------------------------

inicio = dt.datetime.now().strftime("%H:%M:%S")

tool = lt.LanguageTool("pt-BR")
print("Conectado ao servidor ...")

REGRAS_VALIDAS = {
'O_FACTO_DA_ACÇÃO',
'LINKING_VERB_PREDICATE_AGREEMENT',
'GENERAL_VERB_AGREEMENT_ERRORS',
'FORMAL_PRA_PARA',
'À_MEDIDA_EM_QUE',
'DIACRITICS',
'LP_PARONYMS',
'PARONYM_CALCULO_606',
'PARONYM_OFICIO_227',
'PARONYM_BENEFICIO_43',
'PHRASAL_VERB_EM',
'A_NIVEL',
'PARONYM_ACIDULA_1',
'VERB_COMMA_CONJUNCTION',
'ALTERNATIVE_CONJUNCTIONS_COMMA',
'CONTRACOES_OBRIGATORIAS',
'REDUNDANT_CONJUNCTIONS',
'GENERAL_PRONOMIAL_COLOCATIONS',
'HAVER'
}


# ------------------------------------------
# FUNÇÕES
# ------------------------------------------

def extrair_paginas_pdf(path_pdf):

    with open(path_pdf, 'rb') as in_file:

        parser = PDFParser(in_file)
        doc = PDFDocument(parser)

        rsrcmgr = PDFResourceManager()

        for page in PDFPage.create_pages(doc):

            output_string = StringIO()

            device = TextConverter(
                rsrcmgr,
                output_string,
                laparams=LAParams()
            )

            interpreter = PDFPageInterpreter(rsrcmgr, device)

            interpreter.process_page(page)

            texto = output_string.getvalue()

            device.close()

            yield texto


def contar_paginas_pdf(path_pdf):

    with open(path_pdf, 'rb') as in_file:

        parser = PDFParser(in_file)
        doc = PDFDocument(parser)

        return doc.catalog['Pages'].resolve()['Count']


def limpar_texto(texto):

    texto = re.sub(r'[\n\r\t]', ' ', texto)
    texto = re.sub(r'[0-9*/%$+\-]', '', texto)

    return texto


def dividir_texto(texto, tamanho=5000):

    return [texto[i:i+tamanho] for i in range(0, len(texto), tamanho)]


def analisar_texto(texto):

    global tool

    matches = []

    for bloco in dividir_texto(texto):

        tentativas = 0

        while tentativas <= 10:

            try:

                time.sleep(0.3)

                resposta = tool.check(bloco)

                matches.extend(resposta)

                break

            except Exception:

                print("Erro de conexão. Reiniciando servidor...")

                time.sleep(2)

                tool = lt.LanguageTool("pt-BR")

                tentativas += 1

    erros = [m for m in matches if m.rule_id in REGRAS_VALIDAS]

    return erros


# ------------------------------------------
# PLANILHAS
# ------------------------------------------

excel1 = Workbook()
data_geral = excel1.active
data_geral.title = "Data_Geral"
data_geral['A1'] = 'Qt_Total'

excel2 = Workbook()
data_local = excel2.active
data_local.title = "Empresas"
data_local["A1"] = "Empresas"
data_local["B1"] = "Qt_Erros"

indice_excel = 1


# ------------------------------------------
# ESCOLHA DOS CAMINHOS
# ------------------------------------------

path = askdirectory(title='Caminho Data salva')
path_resultado = askdirectory(title='Caminho onde salvar resultados')
path_saida_original = askdirectory(title='Caminho de saída')
path_texto = askdirectory(title='Caminho texto')

lista_caminho = os.listdir(path)

erro_geral = 0
contador_geral = 0


# ------------------------------------------
# LOOP PRINCIPAL
# ------------------------------------------

for empresa in lista_caminho:

    print("*-" * 24)
    print("Empresa:", empresa)

    path_empresa = os.path.join(path, empresa)

    arquivos = os.listdir(path_empresa)

    erro_local = 0

    if len(arquivos) != 0:

        arquivos.sort()

        excel3 = Workbook()
        data_local_ano = excel3.active

        indice_excel1 = 1

        for arquivo in arquivos:

            print("Analisando:", arquivo)

            path_pdf = os.path.join(path_empresa, arquivo)

            num_pages = contar_paginas_pdf(path_pdf)

            texto_completo = ""
            erros_pdf = []

            for pagina in extrair_paginas_pdf(path_pdf):

                pagina = limpar_texto(pagina)

                texto_completo += pagina

                erros_pagina = analisar_texto(pagina)

                erros_pdf.extend(erros_pagina)

            nome_txt = arquivo.replace(".pdf", ".txt")

            path_texto_original = os.path.join(path_texto, nome_txt)

            with open(path_texto_original, 'w', encoding='utf-8') as f:
                f.write(texto_completo)

            erro_anual = len(erros_pdf)

            erro_local += erro_anual

            relatorio = os.path.join(path_resultado, "Erro_" + nome_txt)

            with open(relatorio, 'w', encoding='utf-8') as aperro:

                for i, erro in enumerate(erros_pdf, 1):

                    aperro.write("Erro: " + str(i) + "\n")
                    aperro.write(str(erro) + "\n")

            data_local_ano["A" + str(indice_excel1)] = arquivo
            data_local_ano["B" + str(indice_excel1)] = erro_anual
            data_local_ano["C" + str(indice_excel1)] = num_pages

            indice_excel1 += 1

        excel3.save(os.path.join(path_resultado, empresa + ".xlsx"))

        indice_excel += 1

        data_local["A" + str(indice_excel)] = empresa
        data_local["B" + str(indice_excel)] = erro_local

        erro_geral += erro_local
        contador_geral += 1

    else:

        print("Empresa sem arquivos")

        indice_excel += 1

        data_local["A" + str(indice_excel)] = empresa
        data_local["B" + str(indice_excel)] = 0

        contador_geral += 1

    path_saida = os.path.join(path_saida_original, empresa)

    os.rename(path_empresa, path_saida)


# ------------------------------------------
# SALVAR RESULTADOS
# ------------------------------------------

excel2.save(os.path.join(path_resultado, "Data_Local.xlsx"))

data = pd.read_excel(os.path.join(path_resultado, "Data_Local.xlsx"))

qt_total = data['Qt_Erros'].sum()

data_geral['B1'] = qt_total

excel1.save(os.path.join(path_resultado, "Data_Geral.xlsx"))

fim = dt.datetime.now().strftime("%H:%M:%S")

print("Consulta finalizada com sucesso")
print("Processo começou:", inicio)
print("Processo terminou:", fim)