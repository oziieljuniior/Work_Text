#Biblioteca suporte para extração de texto, reconhecimento do texto como string
from io import StringIO

import time

#Biblioteca para extração de texto
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfparser import PDFParser


#Biblioteca que realiza análise textual
import language_tool_python as lt

#Bibliotecas que configuram os caminhos pelo pc
import os
from tkinter.filedialog import askdirectory

#Blioteca que gerencia planilha
from openpyxl import Workbook
import pandas as pd 


import datetime as dt
inicio = dt.datetime.now().strftime("%H:%M:%S")

#2 planilhas foram criadas: Geral e Local
excel1 = Workbook()
data_geral =  excel1.active
data_geral.title = "Data_Geral"
data_geral['A1'] = 'Qt_Total'
excel2 = Workbook()
data_local = excel2.active
data_local.title = "Empresas"
data_local["A1"] = "Empresas"
data_local["B1"] = "Qt_Erros"
data_local["C1"] = "Qt_Paginas"
indice_excel = 1


#Configuração da biblioteca para análise textual 
tool = lt.LanguageTool("pt-BR", config = {'requestLimit' : 10, 'timeoutRequestLimit': 60})
print("Conectado ao servidor ...")
        


#Configurar caminho
path = askdirectory(title = 'Caminho Data salva')
path_resultado = askdirectory(title = 'Caminho onde o resultado da pesquisa deve ser salvo')
path_saida_original = askdirectory(title = 'Caminho de Saida')
path_texto = askdirectory(title = 'Caminho Texto')
lista_caminho = os.listdir(path)
print(lista_caminho)

#Contadores iniciais para análise de erro
erro_geral = 0
contador_geral = 0

#Percorre e lista arquivos na pasta geral
#caminho é a empresa
for caminho in lista_caminho:
    #configuração para acompanhamento no terminal
    print("*-" * 24)
    print("Fase 1")
    print("Total: ", len(lista_caminho))
    print("Do total estamos na pasta: ", contador_geral + 1)
    print("Analisando: ", caminho)
    
    #Configuração de caminho para percorrer pastas das empresas
    path1 = path + "/" + caminho
    caminho_pdf = os.listdir(path1)

    #Contador dos erros locais de cada empresa
    erro_local = 0

    #Condicional para realizar a captura e análise textual
    if len(caminho_pdf) != 0:
        caminho_pdf.sort()
        #contador qualquer
        contador = 0
        excel3 = Workbook()
        data_local_ano = excel3.active
        data_local_ano.title = "Empresas_Anos"
        indice_excel1 = 1

        for caminho1 in caminho_pdf:
            #configuração para acomponhar análise pelo terminal
            print("*-" * 24)
            print("Fase 2 - Analisando ", len(caminho_pdf))
            print("Arquivo atual: ", caminho1)
            print("Contador: ", contador + 1)
            print((round((contador+1)/len(caminho_pdf),2))*100,"%")
            
            #Ajuste para criação de relatório local de cada empresa e ano
            titulo = caminho1.replace(".pdf",".txt")
            path_relatorio = path_resultado + "/Erro_" + titulo
            print(path_relatorio)
            
            #configurando texto txt
            path_texto_original = path_texto +"/"+ titulo
            print(path_texto_original)
            texto_txt = open(path_texto_original, 'x', encoding = 'utf-8')
            
            #Configuração para extração do texto
            path2 = path1 + "/" + caminho1
            output_string = StringIO()
#                        ...
            with open(path2,'rb') as in_file:
                parser = PDFParser(in_file)
                doc = PDFDocument(parser)
                #numero de paginas
                num_pages = doc.catalog['Pages'].resolve()['Count']
                print(f'O arquivo PDF tem {num_pages} páginas.')
                rsrcmgr = PDFResourceManager()
                device = TextConverter(rsrcmgr, output_string, laparams=LAParams())
                interpreter = PDFPageInterpreter(rsrcmgr,device)
                page_count = 0
                for page in PDFPage.create_pages(doc):
                    if page_count < 50:
                        interpreter.process_page(page)
                        page_count += 1
                    else:
                        break
#            ...
            #variavel de retorno da extração do texto
            text = output_string.getvalue()
            bad = ['\n', '\r','\t','\n\x0c','1','2','3','4','5','6','7','8','9','0','*','-','/','%','$','(.)','(..)','(...)','()','( )','(   )',',.','..','+']
            for name in bad:
                text = text.replace(name, '')
            texto_txt.write(text)
            print("Extração de texto completa!")
            texto_txt.close()
            
            

            #Análise textual
            print("Analisando erros")
            print(dt.datetime.now().strftime("%H:%M:%S"))
            z = 0

            while z <= 20:
                try:
                    time.sleep(3)
                    matches = tool.check(text)
                    print(len(matches))
                            
                    z = 21
                except lt.utils.LanguageToolConfig:
                   print('Error de conexão. Tentando novamente ...')
                   time.sleep(5)
                   tool = lt.LanguageTool("pt-BR", config = {'requestLimit' : 10, 'timeoutRequestLimit': 60})
                   z += 1
                   continue

            print('Procurando erros ... ')
            erros_apontados = len(matches)
            #criação de arquivo txt para relatório local 
            aperro = open(path_relatorio, 'x', encoding='utf-8')
            
            erro_anual = 0
            
            erros = [match for match in matches if match.rule_id == 'O_FACTO_DA_ACÇÃO' or match.rule_id == 'LINKING_VERB_PREDICATE_AGREEMENT' or match.rule_id == 'GENERAL_VERB_AGREEMENT_ERRORS' or match.rule_id == 'FORMAL_PRA_PARA' or match.rule_id == 'À_MEDIDA_EM_QUE' or match.rule_id == 'DIACRITICS' or match.rule_id == 'LP_PARONYMS' or match.rule_id == 'PARONYM_CALCULO_606' or match.rule_id == 'PARONYM_OFICIO_227' or match.rule_id == 'PARONYM_BENEFICIO_43' or match.rule_id == 'PHRASAL_VERB_EM' or match.rule_id == 'A_NIVEL' or match.rule_id == 'PARONYM_ACIDULA_1' or match.rule_id == 'VERB_COMMA_CONJUNCTION' or match.rule_id == 'ALTERNATIVE_CONJUNCTIONS_COMMA' or match.rule_id == 'CONTRACOES_OBRIGATORIAS' or match.rule_id == 'REDUNDANT_CONJUNCTIONS' or match.rule_id == 'GENERAL_PRONOMIAL_COLOCATIONS' or match.rule_id == 'HAVER' or match.rule_id == 'FRAGMENT_TWO_ARTICLES' or match.rule_id == 'PHRASAL_VERB_A' or match.rule_id == 'REFLEXIVE_VERB_SE_AGREEMENT' or match.rule_id == 'PARONYM_INICIO_169' or match.rule_id == 'ATRAVES_DE_POR' or match.rule_id == 'TODOS_FOLLEWED_BY_NOUN_PLURAL' or match.rule_id == 'CRASE_CONFUSION_2' or match.rule_id == 'SIMPLIFICAR_QUE_E_TEM_TÊM' or match.rule_id == 'CRASE_CONFUSION' or match.rule_id == 'CHAMAR_DENOMINAR_DE' or match.rule_id == 'SIMPLIFICAR_CONVERTER_PARA_VERBO_INFINITIVO' or match.rule_id == 'E_QUE_VERBO_E_VERBO' or match.rule_id == 'SIMPLIFICAR_QUE_E_TEM_TÊM' or match.rule_id == 'VERB_QUE_É_VERB_SER' or match.rule_id == 'REDUNDANCY_REPLACE' or match.rule_id == 'QUE_É-SÃO_NC-ADJ_COMO-POR' or match.rule_id == 'QUE_TER_COMO_POR_NOME_ADJ' or match.rule_id == 'DIFERENTES' or match.rule_id == 'QUE_VERBO_A_VERBOINFINITIVO' or match.rule_id == 'REDUNDANCY_PAÍSES' or match.rule_id == 'REDUNDANCY_8_AUMENTAR' or match.rule_id == 'QUE_SUBJ_VS_INF_PESS' or match.rule_id == 'QUE_HÁ_QUE_NÃO_HÁ' or match.rule_id == 'ESTAR_CLARO_DE_QUE' or match.rule_id == 'DEPOIS_DE_APÓS' or match.rule_id == 'SER_CAPAZ_DE_CONSEGUIR' or match.rule_id == 'SIMPLIFICAR_O_QUE_VERBO_VERBOGERUNDIO' or match.rule_id == 'QUE_FORAM_FOI_SÃO_É_SENDO' or match.rule_id == 'QUE_FORAM_FOI_SÃO_É_SENDO' or match.rule_id == 'QUE_ESTAR_CONTRACAO_PREPOSICAO' or match.rule_id == 'VIR_A_VERBO_VERBO' or match.rule_id == 'QUANDO_POSSA_SER' or match.rule_id == 'REDUNDANCY_27_EXPRESSAMENTE' or match.rule_id == 'TER_PARTICIPIO-PASSADO' or match.rule_id == 'REDUNDANCY_28_PERMANECER' or match.rule_id == 'CUJO_LIGACAO_NOME_ADJETIVO_NUMERAL' or match.rule_id == 'AULAS-ENSINO_A_DISTANCIA' or match.rule_id == 'GRAMMATICAL_DOUBLE_NEGATIVES' or match.rule_id == 'PARONYM_ARVORE_0' or match.rule_id == 'PARONYM_CONSORCIO_70' or match.rule_id == 'ETC_USAGE' or match.rule_id == 'PHRASAL_VERB_RESIDIR_EM' or match.rule_id == 'REDUNDANCY_34_JUNTAMENTE' or match.rule_id =='PARONYM_AGENCIA_9' or match.rule_id == 'VÍDEO_CONFERÊNCIA' or match.rule_id == 'SOB_O_PONTO_DE_VISTA' or match.rule_id == 'WORDINESS']
            #for i in range(0, erros_apontados):
                #PORTUGUESE_WORD_REPEAT_BEGINNING_RULE -> Verificar possibilidade de retirar números do relatório or match.rule_id == 'FRAGMENT_TWO_PREPOSITIONS'

            #    if match.rule_id == 'O_FACTO_DA_ACÇÃO' or match.rule_id == 'LINKING_VERB_PREDICATE_AGREEMENT' or match.rule_id == 'GENERAL_VERB_AGREEMENT_ERRORS' or match.rule_id == 'FORMAL_PRA_PARA' or match.rule_id == 'À_MEDIDA_EM_QUE' or match.rule_id == 'DIACRITICS' or match.rule_id == 'LP_PARONYMS' or match.rule_id == 'PARONYM_CALCULO_606' or match.rule_id == 'PARONYM_OFICIO_227' or match.rule_id == 'PARONYM_BENEFICIO_43' or match.rule_id == 'PHRASAL_VERB_EM' or match.rule_id == 'A_NIVEL' or match.rule_id == 'PARONYM_ACIDULA_1' or match.rule_id == 'VERB_COMMA_CONJUNCTION' or match.rule_id == 'ALTERNATIVE_CONJUNCTIONS_COMMA' or match.rule_id == 'CONTRACOES_OBRIGATORIAS' or match.rule_id == 'REDUNDANT_CONJUNCTIONS' or match.rule_id == 'GENERAL_PRONOMIAL_COLOCATIONS' or match.rule_id == 'HAVER' or match.rule_id == 'FRAGMENT_TWO_ARTICLES' or match.rule_id == 'PHRASAL_VERB_A' or match.rule_id == 'REFLEXIVE_VERB_SE_AGREEMENT' or match.rule_id == 'PARONYM_INICIO_169' or match.rule_id == 'ATRAVES_DE_POR' or match.rule_id == 'TODOS_FOLLEWED_BY_NOUN_PLURAL' or match.rule_id == 'CRASE_CONFUSION_2' or match.rule_id == 'SIMPLIFICAR_QUE_E_TEM_TÊM' or match.rule_id == 'CRASE_CONFUSION' or match.rule_id == 'CHAMAR_DENOMINAR_DE' or match.rule_id == 'SIMPLIFICAR_CONVERTER_PARA_VERBO_INFINITIVO' or match.rule_id == 'E_QUE_VERBO_E_VERBO' or match.rule_id == 'SIMPLIFICAR_QUE_E_TEM_TÊM' or match.rule_id == 'VERB_QUE_É_VERB_SER' or match.rule_id == 'REDUNDANCY_REPLACE' or match.rule_id == 'QUE_É-SÃO_NC-ADJ_COMO-POR' or match.rule_id == 'QUE_TER_COMO_POR_NOME_ADJ' or match.rule_id == 'DIFERENTES' or match.rule_id == 'QUE_VERBO_A_VERBOINFINITIVO' or match.rule_id == 'REDUNDANCY_PAÍSES' or match.rule_id == 'REDUNDANCY_8_AUMENTAR' or match.rule_id == 'QUE_SUBJ_VS_INF_PESS' or match.rule_id == 'QUE_HÁ_QUE_NÃO_HÁ' or match.rule_id == 'ESTAR_CLARO_DE_QUE' or match.rule_id == 'DEPOIS_DE_APÓS' or match.rule_id == 'SER_CAPAZ_DE_CONSEGUIR' or match.rule_id == 'SIMPLIFICAR_O_QUE_VERBO_VERBOGERUNDIO' or match.rule_id == 'QUE_FORAM_FOI_SÃO_É_SENDO' or match.rule_id == 'QUE_FORAM_FOI_SÃO_É_SENDO' or match.rule_id == 'QUE_ESTAR_CONTRACAO_PREPOSICAO' or match.rule_id == 'VIR_A_VERBO_VERBO' or match.rule_id == 'QUANDO_POSSA_SER' or match.rule_id == 'REDUNDANCY_27_EXPRESSAMENTE' or match.rule_id == 'TER_PARTICIPIO-PASSADO' or match.rule_id == 'REDUNDANCY_28_PERMANECER' or match.rule_id == 'CUJO_LIGACAO_NOME_ADJETIVO_NUMERAL' or match.rule_id == 'AULAS-ENSINO_A_DISTANCIA' or match.rule_id == 'GRAMMATICAL_DOUBLE_NEGATIVES' or match.rule_id == 'PARONYM_ARVORE_0' or match.rule_id == 'PARONYM_CONSORCIO_70' or match.rule_id == 'ETC_USAGE' or match.rule_id == 'PHRASAL_VERB_RESIDIR_EM' or match.rule_id == 'REDUNDANCY_34_JUNTAMENTE' or match.rule_id =='PARONYM_AGENCIA_9' or match.rule_id == 'VÍDEO_CONFERÊNCIA' or match.rule_id == 'SOB_O_PONTO_DE_VISTA':
            for error in erros:
                    erro_local = erro_local + 1
                    erro_anual = erro_anual + 1
                    #print(match)
                    aperro.write("Erro: ")
                    aperro.write(str(erro_local))
                    aperro.write('\n')
                    aperro.write(str(error))
                    aperro.write('\n')
            aperro.close()
            
            indice_planilhaeA = "A" + str(indice_excel1)
            indice_planilhaeB = "B" + str(indice_excel1)
            indice_planilhaeC = 'C' + str(indice_excel1)
            data_local_ano[indice_planilhaeA] = caminho1
            data_local_ano[indice_planilhaeB] = erro_anual
            data_local_ano[indice_planilhaeC] = num_pages
            indice_excel1 += 1
            
            print("Análise realizada com sucesso!")
            print("Relatório criado com sucesso")
            
            contador += 1
        
        name = path_resultado + "/" + caminho + ".xlsx"
        print("Salvo: ", name)
        excel3.save(name)
        
        indice_excel += 1
        indice_planilhaA = 'A' + str(indice_excel)
        indice_planilhaB = 'B' + str(indice_excel)
        data_local[indice_planilhaA] = caminho
        data_local[indice_planilhaB] = erro_local
        print("Planilha local atualizada")
        print("Quantidade de erros localizadas: ", erro_local)

       

        erro_geral = erro_geral + erro_local
        contador_geral += 1

    
    else:
        print("*-" * 24)
        print("Fase 2 - Analisando ", len(caminho_pdf))
        print(caminho, "não contém arquivo(s)")
        print("Do total estamos na pasta: ", contador_geral + 1)
        
        indice_excel += 1
        indice_planilhaA = 'A' + str(indice_excel)
        indice_planilhaB = 'B' + str(indice_excel)
        data_local[indice_planilhaA] = caminho
        data_local[indice_planilhaB] = erro_local
        
        erro_geral = erro_geral + erro_local
        contador_geral += 1
    
    print("Realizando ajuste de pasta: ", path1)
    path_saida = path_saida_original + "/" + caminho
    print("Para: ", path_saida)
    os.rename(path1, path_saida)


excel2.save(path_resultado + "/Data_Local.xlsx")
print("Calculando quantidade total de erros")
a = []
data = pd.read_excel(path_resultado + "/Data_Local.xlsx")
#print(len(data))
#print(data.columns)
for i in range(0,len(data)):
    print(data['Qt_Erros'][i])
    a.append(data['Qt_Erros'][i])
qt_total = sum(a)
data_geral['B1'] = qt_total

excel1.save(path_resultado + "/Data_Geral.xlsx")

fim = dt.datetime.now().strftime("%H:%M:%S")


print("Consulta finalizada com sucesso")
print("Processo começou: ", inicio)
print("Processo finalizou: ", fim)