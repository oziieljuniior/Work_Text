#Biblioteca suporte para extração de texto, reconhecimento do texto como string
from io import StringIO

import time

#Biblioteca para extração de texto
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfparser import PDFParser

#Biblioteca que realiza análise textual
import language_tool_python as lt

#Bibliotecas que configuram os caminhos pelo pc
import os
from tkinter.filedialog import askdirectory

#Blioteca que gerencia planilha
from openpyxl import Workbook
import pandas as pd 


import datetime as dt
inicio = dt.datetime.now().strftime("%H:%M:%S")

#2 planilhas foram criadas: Geral e Local
excel1 = Workbook()
data_geral =  excel1.active
data_geral.title = "Data_Geral"
data_geral['A1'] = 'Qt_Total'
excel2 = Workbook()
data_local = excel2.active
data_local.title = "Empresas"
data_local["A1"] = "Empresas"
data_local["B1"] = "Qt_Erros"
indice_excel = 1


#Configuração da biblioteca para análise textual 
#tool = lt.LanguageTool("pt-BR")
        


#Configurar caminho
path = askdirectory(title = 'Caminho Data salva')
path_resultado = askdirectory(title = 'Caminho onde o resultado da pesquisa deve ser salvo')
path_saida_original = askdirectory(title = 'Caminho de Saida')
path_texto = askdirectory(title = 'Caminho Texto')
#path.replace("/","\\")
#print(path)
#'/home/oziel/Documentos/Alunos/PauloB/data/today'
#path = 'C:\\Users\\Riallen\\Documents\\Att\\treinamento5 '
lista_caminho = os.listdir(path)
print(lista_caminho)

#Contadores iniciais para análise de erro
erro_geral = 0
contador_geral = 0

#Percorre e lista arquivos na pasta geral
#caminho é a empresa
for caminho in lista_caminho:
    #configuração para acompanhamento no terminal
    print("*-" * 24)
    print("Fase 1")
    print("Total: ", len(lista_caminho))
    print("Do total estamos na pasta: ", contador_geral + 1)
    print("Analisando: ", caminho)
    
    #Configuração de caminho para percorrer pastas das empresas
    path1 = path + "/" + caminho
    caminho_pdf = os.listdir(path1)
    #print(caminho_pdf)

    #Contador dos erros locais de cada empresa
    erro_local = 0

    tool = lt.LanguageTool("pt-BR", config = {'requestLimit' : 10, 'timeoutRequestLimit': 60})
    print("Conectado ao servidor ...")
    
    #Condicional para realizar a captura e análise textual
    if len(caminho_pdf) != 0:
        caminho_pdf.sort()
        #contador qualquer
        contador = 0
        excel3 = Workbook()
        data_local_ano = excel3.active
        data_local_ano.title = "Empresas_Anos"
        indice_excel1 = 1

        for caminho1 in caminho_pdf:
            #configuração para acomponhar análise pelo terminal
            print("*-" * 24)
            print("Fase 2 - Analisando ", len(caminho_pdf))
            print("Arquivo atual: ", caminho1)
            print("Contador: ", contador + 1)
            print((round((contador+1)/len(caminho_pdf),2))*100,"%")
            
            #Ajuste para criação de relatório local de cada empresa e ano
            titulo = caminho1.replace(".pdf",".txt")
            path_relatorio = path_resultado + "/Erro_" + titulo
            print(path_relatorio)
            
            #configurando texto txt
            path_texto_original = path_texto +"/"+ titulo
            print(path_texto_original)
            texto_txt = open(path_texto_original, 'x', encoding = 'utf-8')
            
            #Configuração para extração do texto
            path2 = path1 + "/" + caminho1
            output_string = StringIO()
            with open(path2,'rb') as in_file:
                parser = PDFParser(in_file)
                doc = PDFDocument(parser)
                rsrcmgr = PDFResourceManager()
                device = TextConverter(rsrcmgr, output_string, laparams=LAParams())
                interpreter = PDFPageInterpreter(rsrcmgr,device)
                for page in PDFPage.create_pages(doc):
                    interpreter.process_page(page)
            #variavel de retorno da extração do texto
            text = output_string.getvalue()
            bad = ['\n', '\r','\t','\n\x0c','1','2','3','4','5','6','7','8','9','0','*','-','/','%','$','(.)','(..)','(...)','()','( )','(   )',',.','..','+']
            for name in bad:
                text = text.replace(name, '')
            texto_txt.write(text)
            print("Extração de texto completa!")
            texto_txt.close()
            
            

            #Análise textual
            print("Analisando erros")
            print(dt.datetime.now().strftime("%H:%M:%S"))
            z = 0

            while z <= 20:
                try:
                    matches = tool.check(text)
                    print(len(matches))
                            
                    z = 21
                except lt.utils.LanguageToolError:
                   print('Error de conexão. Tentando novamente ...')
                   time.sleep(5)
                   tool = lt.LanguageTool("pt-BR", config = {'requestLimit' : 10, 'timeoutRequestLimit': 60})
                   z += 1
                   continue

            print('Procurando erros ... ')
            erros_apontados = len(matches)
            #criação de arquivo txt para relatório local 
            aperro = open(path_relatorio, 'x', encoding='utf-8')
            
            erro_anual = 0
            
            for i in range(0, erros_apontados):
                #PORTUGUESE_WORD_REPEAT_BEGINNING_RULE -> Verificar possibilidade de retirar números do relatório or matches[i].ruleId == 'FRAGMENT_TWO_PREPOSITIONS'
                if matches[i].ruleId == 'O_FACTO_DA_ACÇÃO' or matches[i].ruleId == 'LINKING_VERB_PREDICATE_AGREEMENT' or matches[i].ruleId == 'GENERAL_VERB_AGREEMENT_ERRORS' or matches[i].ruleId == 'FORMAL_PRA_PARA' or matches[i].ruleId == 'À_MEDIDA_EM_QUE' or matches[i].ruleId == 'DIACRITICS' or matches[i].ruleId == 'LP_PARONYMS' or matches[i].ruleId == 'PARONYM_CALCULO_606' or matches[i].ruleId == 'PARONYM_OFICIO_227' or matches[i].ruleId == 'PARONYM_BENEFICIO_43' or matches[i].ruleId == 'PHRASAL_VERB_EM' or matches[i].ruleId == 'A_NIVEL' or matches[i].ruleId == 'PARONYM_ACIDULA_1' or matches[i].ruleId == 'VERB_COMMA_CONJUNCTION' or matches[i].ruleId == 'ALTERNATIVE_CONJUNCTIONS_COMMA' or matches[i].ruleId == 'CONTRACOES_OBRIGATORIAS' or matches[i].ruleId == 'REDUNDANT_CONJUNCTIONS' or matches[i].ruleId == 'GENERAL_PRONOMIAL_COLOCATIONS' or matches[i].ruleId == 'HAVER' or matches[i].ruleId == 'FRAGMENT_TWO_ARTICLES' or matches[i].ruleId == 'PHRASAL_VERB_A' or matches[i].ruleId == 'REFLEXIVE_VERB_SE_AGREEMENT' or matches[i].ruleId == 'PARONYM_INICIO_169' or matches[i].ruleId == 'ATRAVES_DE_POR' or matches[i].ruleId == 'TODOS_FOLLEWED_BY_NOUN_PLURAL' or matches[i].ruleId == 'CRASE_CONFUSION_2' or matches[i].ruleId == 'SIMPLIFICAR_QUE_E_TEM_TÊM' or matches[i].ruleId == 'CRASE_CONFUSION' or matches[i].ruleId == 'CHAMAR_DENOMINAR_DE' or matches[i].ruleId == 'SIMPLIFICAR_CONVERTER_PARA_VERBO_INFINITIVO' or matches[i].ruleId == 'E_QUE_VERBO_E_VERBO' or matches[i].ruleId == 'SIMPLIFICAR_QUE_E_TEM_TÊM' or matches[i].ruleId == 'VERB_QUE_É_VERB_SER' or matches[i].ruleId == 'REDUNDANCY_REPLACE' or matches[i].ruleId == 'QUE_É-SÃO_NC-ADJ_COMO-POR' or matches[i].ruleId == 'QUE_TER_COMO_POR_NOME_ADJ' or matches[i].ruleId == 'DIFERENTES' or matches[i].ruleId == 'QUE_VERBO_A_VERBOINFINITIVO' or matches[i].ruleId == 'REDUNDANCY_PAÍSES' or matches[i].ruleId == 'REDUNDANCY_8_AUMENTAR' or matches[i].ruleId == 'QUE_SUBJ_VS_INF_PESS' or matches[i].ruleId == 'QUE_HÁ_QUE_NÃO_HÁ' or matches[i].ruleId == 'ESTAR_CLARO_DE_QUE' or matches[i].ruleId == 'DEPOIS_DE_APÓS' or matches[i].ruleId == 'SER_CAPAZ_DE_CONSEGUIR' or matches[i].ruleId == 'SIMPLIFICAR_O_QUE_VERBO_VERBOGERUNDIO' or matches[i].ruleId == 'QUE_FORAM_FOI_SÃO_É_SENDO' or matches[i].ruleId == 'QUE_FORAM_FOI_SÃO_É_SENDO' or matches[i].ruleId == 'QUE_ESTAR_CONTRACAO_PREPOSICAO' or matches[i].ruleId == 'VIR_A_VERBO_VERBO' or matches[i].ruleId == 'QUANDO_POSSA_SER' or matches[i].ruleId == 'REDUNDANCY_27_EXPRESSAMENTE' or matches[i].ruleId == 'TER_PARTICIPIO-PASSADO' or matches[i].ruleId == 'REDUNDANCY_28_PERMANECER' or matches[i].ruleId == 'CUJO_LIGACAO_NOME_ADJETIVO_NUMERAL' or matches[i].ruleId == 'AULAS-ENSINO_A_DISTANCIA' or matches[i].ruleId == 'GRAMMATICAL_DOUBLE_NEGATIVES' or matches[i].ruleId == 'PARONYM_ARVORE_0' or matches[i].ruleId == 'PARONYM_CONSORCIO_70' or matches[i].ruleId == 'ETC_USAGE' or matches[i].ruleId == 'PHRASAL_VERB_RESIDIR_EM' or matches[i].ruleId == 'REDUNDANCY_34_JUNTAMENTE' or matches[i].ruleId =='PARONYM_AGENCIA_9' or matches[i].ruleId == 'VÍDEO_CONFERÊNCIA' or matches[i].ruleId == 'SOB_O_PONTO_DE_VISTA':
                    erro_local = erro_local + 1
                    erro_anual = erro_anual + 1
                    #print(matches[i])
                    aperro.write("Erro: ")
                    aperro.write(str(erro_local))
                    aperro.write('\n')
                    aperro.write(str(matches[i]))
                    aperro.write('\n')
            aperro.close()
            
            indice_planilhaeA = "A" + str(indice_excel1)
            indice_planilhaeB = "B" + str(indice_excel1)
            data_local_ano[indice_planilhaeA] = caminho1
            data_local_ano[indice_planilhaeB] = erro_anual
            indice_excel1 += 1
            
            print("Análise realizada com sucesso!")
            print("Relatório criado com sucesso")
            
            contador += 1
        
        name = path_resultado + "/" + caminho + ".xlsx"
        print("Salvo: ", name)
        excel3.save(name)
        
        indice_excel += 1
        indice_planilhaA = 'A' + str(indice_excel)
        indice_planilhaB = 'B' + str(indice_excel)
        data_local[indice_planilhaA] = caminho
        data_local[indice_planilhaB] = erro_local
        print("Planilha local atualizada")
        print("Quantidade de erros localizadas: ", erro_local)

       

        erro_geral = erro_geral + erro_local
        contador_geral += 1

    
    else:
        print("*-" * 24)
        print("Fase 2 - Analisando ", len(caminho_pdf))
        print(caminho, "não contém arquivo(s)")
        print("Do total estamos na pasta: ", contador_geral + 1)
        
        indice_excel += 1
        indice_planilhaA = 'A' + str(indice_excel)
        indice_planilhaB = 'B' + str(indice_excel)
        data_local[indice_planilhaA] = caminho
        data_local[indice_planilhaB] = erro_local
        
        erro_geral = erro_geral + erro_local
        contador_geral += 1
    
    print("Realizando ajuste de pasta: ", path1)
    path_saida = path_saida_original + "/" + caminho
    print("Para: ", path_saida)
    os.rename(path1, path_saida)


excel2.save(path_resultado + "/Data_Local.xlsx")
print("Calculando quantidade total de erros")
a = []
data = pd.read_excel(path_resultado + "/Data_Local.xlsx")
#print(len(data))
#print(data.columns)
for i in range(0,len(data)):
    print(data['Qt_Erros'][i])
    a.append(data['Qt_Erros'][i])
qt_total = sum(a)
data_geral['B1'] = qt_total

excel1.save(path_resultado + "/Data_Geral.xlsx")

fim = dt.datetime.now().strftime("%H:%M:%S")


print("Consulta finalizada com sucesso")
print("Processo começou: ", inicio)
print("Processo finalizou: ", fim)