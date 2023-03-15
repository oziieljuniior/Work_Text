import os 
import pandas as pd
from openpyxl import Workbook

#path_geral = 'C:\\Users\\Riallen\\Documents\\Work_Text\\Treinamento_Geral'
#pasta_inicial = os.listdir(path_geral)
#t0 = len(pasta_inicial)

#i = 1
#while i <= t0:
#    print(i)
#    path_local = path_geral + '\\Treinamento' + str(i) + '\\Resultado_Treinamento' + str(i) 
    #rename_antes1 = path_local + '\\Data_Geral.xlsx'
    #rename_antes2 = path_local + '\\Data_Local.xlsx'
    #rename1 = 'C:\\Users\\Riallen\\Documents\\Work_Text\\Erros\\Data_Geral' + str(i) + '.xlsx' 
    #rename2 = 'C:\\Users\\Riallen\\Documents\\Work_Text\\Erros\\Data_Local' + str(i) + '.xlsx' 
    #os.rename(rename_antes1,rename1)
    #os.rename(rename_antes2,rename2)
#    i += 1

path_geral = 'C:\\Users\\Riallen\\Documents\\Work_Text\\Erros\\Data_Geral'
path_local = 'C:\\Users\\Riallen\\Documents\\Work_Text\\Erros\\Data_Local'

pasta_geral = os.listdir(path_geral)
pasta_local = os.listdir(path_local)
t0 = len(pasta_geral)
#print(pasta_local)
#print(pasta_geral)
i = 1
#ativar = Workbook()
#tabela = ativar.active
#tabela.title = 'Qt_Total'
#tabela['A1'] = 'Qt_Total'
#tabela['B1'] = 'Numero_Erros'
#indice_excel = 2

#ativar2 = Workbook()
#tabela2 = ativar2.active
#tabela2.title = 'Qt_Total'
#tabela2['A1'] = 'Qt_Total'
#tabela2['B1'] = 'Numeros_Erros'

#soma = 0

#while i <= t0:
#    name1 = '\\Data_Geral' + str(i) + '.xlsx'
#    path_data = path_geral + name1
#    data1 = pd.read_excel(path_data, header = None)
#    print(data1[0],data1[1])

#    arr = 'A' + str(indice_excel)
#    brr = 'B' + str(indice_excel)
#    tabela[arr] = 'Qt_Total' + str(i)
#    tabela[brr] = int(data1[1])

#    soma = data1[1] + soma

#    i += 1
#    indice_excel += 1

#ativar.save('C:\\Users\\Riallen\\Documents\\Work_Text\\Erros\\Erro_Geral_Acumulado.xlsx')

#tabela2['A2'] = "Soma_Total"
#tabela2['B2'] = int(soma)
#ativar2.save('C:\\Users\\Riallen\\Documents\\Work_Text\\Erros\\Erro_Geral.xlsx')

ativar = Workbook()
tabela = ativar.active
tabela.title = 'Qt_Erro_Local'
tabela['A1'] = 'Empresas'
tabela['B1'] = 'Qt_Erros'
soma = 0
indice_excel = 2


while i <= t0:
    print(i)
    name1 = '\\Data_Local' + str(i) + '.xlsx'
    path_data = path_local + name1
    data1 = pd.read_excel(path_data)
    print(data1)
    t1 = len(data1['Empresas'])
    print(t1)
    j = 0
    while j < t1:
        arr = 'A' + str(indice_excel)
        brr = 'B' + str(indice_excel)
        tabela[arr] = data1['Empresas'][j]
        tabela[brr] = int(data1['Qt_Erros'][j])

        soma = soma + data1['Qt_Erros'][j]

        indice_excel += 1
        j += 1
    
    i += 1
print(soma)
ativar.save('C:\\Users\\Riallen\\Documents\\Work_Text\\Erros\\Erro_Local.xlsx')
