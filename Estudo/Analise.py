from io import StringIO
import os
import time
import re
import datetime as dt

import pandas as pd
import language_tool_python as lt

from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfparser import PDFParser

from tkinter.filedialog import askdirectory
from openpyxl import Workbook


# ------------------------------------------
# CONFIGURAÇÃO INICIAL
# ------------------------------------------

inicio = dt.datetime.now().strftime("%H:%M:%S")

tool = lt.LanguageTool("pt-BR")
print("Conectado ao servidor ...")

REGRAS_VALIDAS = {
'O_FACTO_DA_ACÇÃO',
'LINKING_VERB_PREDICATE_AGREEMENT',
'GENERAL_VERB_AGREEMENT_ERRORS',
'FORMAL_PRA_PARA',
'À_MEDIDA_EM_QUE',
'DIACRITICS',
'LP_PARONYMS',
'PARONYM_CALCULO_606',
'PARONYM_OFICIO_227',
'PARONYM_BENEFICIO_43',
'PHRASAL_VERB_EM',
'A_NIVEL',
'PARONYM_ACIDULA_1',
'VERB_COMMA_CONJUNCTION',
'ALTERNATIVE_CONJUNCTIONS_COMMA',
'CONTRACOES_OBRIGATORIAS',
'REDUNDANT_CONJUNCTIONS',
'GENERAL_PRONOMIAL_COLOCATIONS',
'HAVER'
}


# ------------------------------------------
# FUNÇÕES
# ------------------------------------------

def extrair_texto_pdf(path_pdf, limite_paginas=50):

    output_string = StringIO()

    with open(path_pdf, 'rb') as in_file:

        parser = PDFParser(in_file)
        doc = PDFDocument(parser)

        rsrcmgr = PDFResourceManager()
        device = TextConverter(rsrcmgr, output_string, laparams=LAParams())
        interpreter = PDFPageInterpreter(rsrcmgr, device)

        for i, page in enumerate(PDFPage.create_pages(doc)):

            if i >= limite_paginas:
                break

            interpreter.process_page(page)

    return output_string.getvalue(), doc.catalog['Pages'].resolve()['Count']


def limpar_texto(texto):

    texto = re.sub(r'[\n\r\t]', ' ', texto)
    texto = re.sub(r'[0-9*/%$+\-]', '', texto)

    return texto


def dividir_texto(texto, tamanho=20000):

    return [texto[i:i+tamanho] for i in range(0, len(texto), tamanho)]


def analisar_texto(texto):

    global tool

    matches = []

    for bloco in dividir_texto(texto):

        z = 0

        while z <= 20:

            try:

                time.sleep(2)

                resposta = tool.check(bloco)

                matches.extend(resposta)

                break

            except lt.LanguageToolError:

                print("Erro de conexão. Reiniciando servidor...")

                time.sleep(5)

                tool = lt.LanguageTool("pt-BR")

                z += 1

    erros = [m for m in matches if m.rule_id in REGRAS_VALIDAS]

    return erros


# ------------------------------------------
# PLANILHAS
# ------------------------------------------

excel1 = Workbook()
data_geral = excel1.active
data_geral.title = "Data_Geral"
data_geral['A1'] = 'Qt_Total'

excel2 = Workbook()
data_local = excel2.active
data_local.title = "Empresas"
data_local["A1"] = "Empresas"
data_local["B1"] = "Qt_Erros"

indice_excel = 1


# ------------------------------------------
# ESCOLHA DOS CAMINHOS
# ------------------------------------------

path = askdirectory(title='Caminho Data salva')
path_resultado = askdirectory(title='Caminho onde salvar resultados')
path_saida_original = askdirectory(title='Caminho de saída')
path_texto = askdirectory(title='Caminho texto')

lista_caminho = os.listdir(path)

erro_geral = 0
contador_geral = 0


# ------------------------------------------
# LOOP PRINCIPAL
# ------------------------------------------

for empresa in lista_caminho:

    print("*-" * 24)
    print("Empresa:", empresa)

    path_empresa = os.path.join(path, empresa)

    arquivos = os.listdir(path_empresa)

    erro_local = 0

    if len(arquivos) != 0:

        arquivos.sort()

        excel3 = Workbook()
        data_local_ano = excel3.active

        indice_excel1 = 1

        for arquivo in arquivos:

            print("Analisando:", arquivo)

            path_pdf = os.path.join(path_empresa, arquivo)

            texto, num_pages = extrair_texto_pdf(path_pdf)

            texto = limpar_texto(texto)

            nome_txt = arquivo.replace(".pdf", ".txt")

            path_texto_original = os.path.join(path_texto, nome_txt)

            with open(path_texto_original, 'w', encoding='utf-8') as f:
                f.write(texto)

            erros = analisar_texto(texto)

            erro_anual = len(erros)

            erro_local += erro_anual

            relatorio = os.path.join(path_resultado, "Erro_" + nome_txt)

            with open(relatorio, 'w', encoding='utf-8') as aperro:

                for i, erro in enumerate(erros, 1):

                    aperro.write("Erro: " + str(i) + "\n")
                    aperro.write(str(erro) + "\n")

            data_local_ano["A" + str(indice_excel1)] = arquivo
            data_local_ano["B" + str(indice_excel1)] = erro_anual
            data_local_ano["C" + str(indice_excel1)] = num_pages

            indice_excel1 += 1

        excel3.save(os.path.join(path_resultado, empresa + ".xlsx"))

        indice_excel += 1

        data_local["A" + str(indice_excel)] = empresa
        data_local["B" + str(indice_excel)] = erro_local

        erro_geral += erro_local
        contador_geral += 1

    else:

        print("Empresa sem arquivos")

        indice_excel += 1

        data_local["A" + str(indice_excel)] = empresa
        data_local["B" + str(indice_excel)] = 0

        contador_geral += 1

    path_saida = os.path.join(path_saida_original, empresa)

    os.rename(path_empresa, path_saida)


# ------------------------------------------
# SALVAR RESULTADOS
# ------------------------------------------

excel2.save(os.path.join(path_resultado, "Data_Local.xlsx"))

data = pd.read_excel(os.path.join(path_resultado, "Data_Local.xlsx"))

qt_total = data['Qt_Erros'].sum()

data_geral['B1'] = qt_total

excel1.save(os.path.join(path_resultado, "Data_Geral.xlsx"))

fim = dt.datetime.now().strftime("%H:%M:%S")

print("Consulta finalizada com sucesso")
print("Processo começou:", inicio)
print("Processo terminou:", fim)