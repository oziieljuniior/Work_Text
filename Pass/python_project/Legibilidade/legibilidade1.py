import os
import re
from tkinter import filedialog
import nltk
from openpyxl import Workbook
from io import StringIO
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfparser import PDFParser

# Configuração das planilhas
workbook = Workbook()
data_geral = workbook.active
data_geral.title = "Data Geral"
data_geral["A1"] = "Empresas"
data_geral["B1"] = "Qt. de letras"
data_geral["C1"] = "Qt. de sílabas"
data_geral["D1"] = "Qt. de palavras"
data_geral["E1"] = "Qt. de pal. complexas"
data_geral["F1"] = "Qt. de Sentenças"
data_geral["G1"] = "Qt. de páginas"

data_local = workbook.create_sheet("Data Local")
data_local["A1"] = "Empresas"
data_local["B1"] = "Qt. de letras"
data_local["C1"] = "Qt. de sílabas"
data_local["D1"] = "Qt. de palavras"
data_local["E1"] = "Qt. de pal. complexas"
data_local["F1"] = "Qt. de Sentenças"
data_local["G1"] = "Qt. de páginas"

def count_letters(text):
    """
    Retorna a quantidade de letras em um texto.
    """
    return len(re.findall(r'\w', text))

def count_complex_words(text):
    """
    Retorna a quantidade de palavras complexas em um texto.
    """
    words = nltk.word_tokenize(text)
    num_complex_words = 0
    for word in words:
        num_syllables = count_syllables(word)
        if num_syllables >= 3:
            num_complex_words += 1
    return num_complex_words

def count_pages(file):
    """
    Retorna a quantidade de páginas em um arquivo PDF.
    """
    parser = PDFParser(file)
    doc = PDFDocument(parser)
    return doc.info[0]["Pages"]

def measure_text(filename):
    """
    Executa as medições de legibilidade em um arquivo de texto.
    Adiciona os resultados nas planilhas correspondentes.
    """
    with open(filename, 'r') as file:
        text = file.read()

    # Medição dos critérios
    num_letters = count_letters(text)
    num_syllables = sum([count_syllables(word) for word in text.split()])
    num_words = count_words(text)
    num_complex_words = count_complex_words(text)
    num_sentences = count_sentences(text)

    # Adição dos resultados na planilha "Data Geral"
    data_geral.append([os.path.basename(filename), num_letters, num_syllables, num_words, num_complex_words, num_sentences, "N/A"])

    # Adição dos resultados na planilha "Data Local"
    with open(filename, 'rb') as file:
        num_pages = count_pages(file)
        data_local.append([os.path.basename(filename), num_letters, num_syllables, num_words, num_complex_words, num_sentences, num_pages])

    # Salvando as planilhas em arquivo
    planilha1.save(os.path.join(janela_saida, "Data Geral.xlsx"))
    planilha2.save(os.path.join(janela_saida, "Data Local.xlsx"))

def main():
    """
    Executa o programa principal.
    """
    # Selecionando os arquivos a serem analisados
    filenames = filedialog.askopenfilenames()

    # Processando cada arquivo
    for filename in filenames:
        # Verificando a extensão do arquivo
        ext = os.path.splitext(filename)[1]
        if ext in ('.txt', '.pdf'):
            # Medindo os critérios de legibilidade do arquivo
            measure_text(filename)
        else:
            print(f"Arquivo {filename} não suportado.")
        
    # Calculando a soma dos critérios na planilha "Data Geral"
    for col_num in range(2, 8):
        total = sum([cell.value for cell in data_geral.columns[col_num][1:]])
        data_geral.cell(row=data_geral.max_row+1, column=col_num, value=total)
    
    # Salvando a planilha "Data Geral" com os totais
    planilha1.save(os.path.join(janela_saida, "Data Geral.xlsx"))

if __name__ == "__main__":
    main()
